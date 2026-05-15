#!/usr/bin/env python3
"""
Import script for the cleaned baligirls xlsx → live site.

Produces three artifacts (always written, regardless of mode):
  - import_run/providers.sql              (DB inserts, wrapped in transaction)
  - import_run/gcs_upload.sh              (gcloud storage cp commands)
  - import_run/summary.txt                (human-readable summary)

Modes
-----
  --mode dry-run   (default)  Write artifacts. Do NOT touch GCS or DB.
  --mode live                 Write artifacts AND execute them in order:
                              1) GCS uploads
                              2) SQL transaction (via psql over ssh)

Scope: targets ONLY baligirls.
  - GCS bucket : gs://gda-ce01-bucket/baligirls/uploads/
  - DB         : ascortbali (on gda-pn01, via sudo -u postgres psql)
  - PM2        : NEVER touches essentialbali / schoolcatering apps.

Conventions used (verified against /var/www/baligirls/ on 2026-05-15):
  - providers.password   = bcrypt hash of phone digits
  - providers.temp_password = phone digits (cleartext)
  - providers.username   = <slug>@email.com
  - providers.slug       = slugify(Name) (already deduped vs DB)
  - providers.url        = /creator/preview/<slug>
  - providers.title      = "<Title-Case Category> <Name> - <City> / <Country>"
  - providers.is_active  = TRUE
  - providers.verified   = FALSE (admin toggles per row later)
  - provider_images.image_id        = IMPORT_<provider_id>_<seq:02d>
  - provider_images.sequence_number = 1..N (max 20 per check constraint)
  - provider_images.image_file      = bare filename (no path)
  - GCS object path                 = baligirls/uploads/<image_file>

Skipped from import:
  - .mp4 files (site is image-only via sharp pipeline)
  - Dead DB columns (bust_size dropped; dick_size, city_part, cell_phone_2,
    telegram_id_2, tour, provides left as NULL)
"""
import argparse
import os
import re
import subprocess
import sys
from pathlib import Path

import bcrypt
import openpyxl

# ──────────────────────────────────────────────────────────────────────
# Constants — none of these should change at run time
# ──────────────────────────────────────────────────────────────────────
HERE          = Path(__file__).resolve().parent
XLSX_FILE     = HERE / "bg_data_cleaned_02.xlsx"
IMAGES_DIR    = HERE / "images_02"   # batch-specific: where the photo files live
AVATAR_FILE   = "avatar-default-lady.png"
OUT_DIR       = HERE / "import_run"

GCS_BUCKET    = "gda-ce01-bucket"
GCS_PREFIX    = "baligirls/uploads"
GCS_CREDS     = "/etc/gda-credentials/gda-viceroy-17373de6d690.json"  # on server

SSH_HOST      = "gda-pn01"
SERVER_STAGE  = "/tmp/baligirls_import"  # where we'll scp local files for upload

# Columns we INSERT into providers (positional). Anything not listed is NULL
# or DB default. We intentionally skip the dead columns (dick_size, city_part,
# cell_phone_2, telegram_id_2, tour, provides) and bust_size (column dropped).
PROVIDERS_COLS = [
    "uuid", "provider_id", "username", "password", "url", "slug",
    "title", "temp_password", "last_seen",
    "model_name", "gender", "age",
    "orientation", "ethnicity", "nationality", "languages",
    "height", "weight", "eyes", "hair_color", "hair_length",
    "pubic_hair", "bust_type", "smoker", "tattoo", "piercing",
    "country", "city", "location",
    "phone_number", "cell_phone", "telegram_id", "wechat_id",
    "services", "available_for", "meeting_with",
    "travel", "escort_type", "notes",
    "is_active", "verified", "created_at",
]

# Listing sort: site does `ORDER BY (image IS NULL) ASC, created_at DESC`.
# Our avatar creators DO have an image (the placeholder), so they'd otherwise
# land at the top. Backdate them so they appear after the 100 existing
# creators (which span 2026-02-08 to 2026-03-07).
CREATED_AT_REAL_PHOTO = "2026-05-16 12:00:00"   # batch 02 — newer than batch 01 (2026-05-15)
CREATED_AT_AVATAR     = "2026-01-01 00:00:00"   # before all existing, appears at bottom

# Hair length / hair color / eyes / etc. enums are already site-canonical in
# the xlsx; we don't re-normalize here.

# ──────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────

def sql_str(v):
    """Quote a value for an SQL literal. NULL for None/empty-string."""
    if v is None:
        return "NULL"
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, (int, float)):
        if isinstance(v, float) and v.is_integer():
            v = int(v)
        return str(v)
    s = str(v)
    if s == "":
        return "NULL"
    # PostgreSQL string literal — escape single quotes by doubling.
    return "'" + s.replace("'", "''") + "'"


def slugify(s):
    s = str(s or "").lower().strip()
    return re.sub(r"[^a-z0-9]+", "-", s).strip("-")[:120]


def title_case(s):
    return " ".join(w.capitalize() for w in str(s or "").split())


def bcrypt_hash(plaintext):
    return bcrypt.hashpw(plaintext.encode(), bcrypt.gensalt(rounds=10)).decode()


def sh_quote(s):
    """Quote a path for a POSIX shell line."""
    return "'" + str(s).replace("'", "'\\''") + "'"

# ──────────────────────────────────────────────────────────────────────
# Build artifacts
# ──────────────────────────────────────────────────────────────────────

def build_artifacts():
    OUT_DIR.mkdir(exist_ok=True)
    sql_path = OUT_DIR / "providers.sql"
    gcs_path = OUT_DIR / "gcs_upload.sh"
    sum_path = OUT_DIR / "summary.txt"

    wb = openpyxl.load_workbook(XLSX_FILE, data_only=True)
    ws = wb.active
    H = {c.value: i for i, c in enumerate(ws[1])}

    # Collect every image file (including avatar) we'll need to upload.
    uploads = {}   # filename -> local path
    avatar_local = HERE / AVATAR_FILE
    if avatar_local.exists():
        uploads[AVATAR_FILE] = avatar_local

    rows = []
    skipped_video = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        d = {col: r[i] for col, i in H.items()}
        rows.append(d)
        for fn in str(d.get("Image Files (info only)") or "").split(";"):
            fn = fn.strip()
            if not fn:
                continue
            if fn.lower().endswith(".mp4"):
                skipped_video.append(fn)
                continue
            local = IMAGES_DIR / fn
            if local.exists():
                uploads[fn] = local
            # If the filename is the avatar, AVATAR_FILE handles it above.

    # ── SQL artifact ──
    sql_lines = []
    sql_lines.append("-- baligirls import — generated by import_baligirls.py")
    sql_lines.append(f"-- {len(rows)} providers, images via per-row INSERTs below")
    sql_lines.append("-- Wrapped in a single transaction; ROLLBACK on any failure.")
    sql_lines.append("BEGIN;")
    sql_lines.append("")

    image_inserts = 0
    for d in rows:
        # Derive secondaries
        slug         = d.get("url", "").rsplit("/", 1)[-1] if d.get("url") else slugify(d.get("Name"))
        if not slug:
            slug = slugify(d.get("Name"))
        username     = f"{slug}@email.com"
        phone_digits = "".join(c for c in str(d.get("Phone Number") or "") if c.isdigit())
        pw_hash      = bcrypt_hash(phone_digits) if phone_digits else bcrypt_hash("changeme")
        cat          = str(d.get("Category") or "freelance").strip()
        title        = f"{title_case(cat)} {d.get('Name')} - {d.get('City')} / {d.get('Country')}"
        uses_avatar  = (str(d.get("Image Files (info only)") or "").strip() == AVATAR_FILE)
        created_at   = CREATED_AT_AVATAR if uses_avatar else CREATED_AT_REAL_PHOTO

        # Build the providers INSERT
        values = {
            "uuid":          d["uuid"],
            "provider_id":   d["ID"],
            "username":      username,
            "password":      pw_hash,
            "url":           d["url"],
            "slug":          slug,
            "title":         title,
            "temp_password": d["temp_password"],
            "last_seen":     d["Last seen online"],
            "model_name":    d["Name"],
            "gender":        d["Gender"],
            "age":           d["Age"],
            "orientation":   d["Orientation"],
            "ethnicity":     d["Ethnicity"],
            "nationality":   d["Nationality"],
            "languages":     d["Languages"],
            "height":        d["Height"],
            "weight":        d["Weight"],
            "eyes":          d["Eyes"],
            "hair_color":    d["Hair color"],
            "hair_length":   d["Hair length"],
            "pubic_hair":    d["Pubic hair"],
            "bust_type":     d["Bust type"],
            "smoker":        d["Smoker"],
            "tattoo":        d["Tattoo"],
            "piercing":      d["Piercing"],
            "country":       d["Country"],
            "city":          d["City"],
            "location":      d["Location"],
            "phone_number":  d["Phone Number"],
            "cell_phone":    d["Cell phone"],
            "telegram_id":   d["Telegram ID"],
            "wechat_id":     d["WeChat ID"],
            "services":      d["Services"],
            "available_for": d["Available for"],
            "meeting_with":  d["Meeting with"],
            "travel":        d["Travel"],
            "escort_type":   cat,
            "notes":         d["About Me"],
            "is_active":     True,
            "verified":      False,
            "created_at":    created_at,
        }

        col_list = ", ".join(PROVIDERS_COLS)
        val_list = ", ".join(sql_str(values[c]) for c in PROVIDERS_COLS)
        sql_lines.append(f"-- creator: {d['Name']}  ({phone_digits})")
        sql_lines.append(f"INSERT INTO providers ({col_list}) VALUES ({val_list});")

        # Photos for this creator
        images = [fn.strip() for fn in str(d.get("Image Files (info only)") or "").split(";") if fn.strip()]
        # Filter mp4
        images = [i for i in images if not i.lower().endswith(".mp4")]
        seq = 0
        for fn in images[:20]:  # respect DB check (1..20)
            seq += 1
            image_id = f"IMPORT_{d['ID']}_{seq:02d}"
            sql_lines.append(
                f"INSERT INTO provider_images (image_id, provider_uuid, provider_id, image_file, sequence_number, resolution) "
                f"VALUES ({sql_str(image_id)}, {sql_str(d['uuid'])}::uuid, {sql_str(d['ID'])}, {sql_str(fn)}, {seq}, 'clean');"
            )
            image_inserts += 1
        sql_lines.append("")

    sql_lines.append("COMMIT;")
    sql_path.write_text("\n".join(sql_lines))

    # ── GCS uploads artifact ──
    gcs_lines = [
        "#!/usr/bin/env bash",
        "# baligirls image upload — gcloud storage cp to GCS",
        "# Run on gda-pn01 (where the credential JSON is mounted).",
        "set -euo pipefail",
        f"export GOOGLE_APPLICATION_CREDENTIALS={GCS_CREDS}",
        f"BUCKET={sh_quote('gs://' + GCS_BUCKET + '/' + GCS_PREFIX)}",
        f"STAGE={sh_quote(SERVER_STAGE)}",
        "",
        f"# {len(uploads)} files to upload",
    ]
    for fn in sorted(uploads):
        gcs_lines.append(f"gcloud storage cp \"$STAGE/{fn}\" \"$BUCKET/{fn}\"")
    gcs_path.write_text("\n".join(gcs_lines) + "\n")
    gcs_path.chmod(0o755)

    # ── Summary ──
    sum_lines = [
        f"Total providers rows to insert: {len(rows)}",
        f"Total provider_images rows:     {image_inserts}",
        f"Total GCS uploads:              {len(uploads)} files",
        f"Skipped (video):                {len(skipped_video)}  -> {skipped_video}",
        "",
        "Artifacts written:",
        f"  {sql_path.relative_to(HERE)}",
        f"  {gcs_path.relative_to(HERE)}",
        f"  {sum_path.relative_to(HERE)}",
    ]
    sum_path.write_text("\n".join(sum_lines) + "\n")
    print("\n".join(sum_lines))
    return uploads, sql_path, gcs_path

# ──────────────────────────────────────────────────────────────────────
# Live execution (only runs when --mode live)
# ──────────────────────────────────────────────────────────────────────

def run_live(uploads, sql_path, gcs_path):
    print("\n[LIVE] 1/3 — scp images to server stage")
    subprocess.run(["ssh", SSH_HOST, f"mkdir -p {SERVER_STAGE}"], check=True)
    # Build list and scp in one batch
    local_files = [str(p) for p in uploads.values()]
    subprocess.run(["scp", "-q", *local_files, f"{SSH_HOST}:{SERVER_STAGE}/"], check=True)

    print("\n[LIVE] 2/3 — scp gcs_upload.sh and run on server")
    subprocess.run(["scp", "-q", str(gcs_path), f"{SSH_HOST}:{SERVER_STAGE}/gcs_upload.sh"], check=True)
    subprocess.run(["ssh", SSH_HOST, f"bash {SERVER_STAGE}/gcs_upload.sh"], check=True)

    print("\n[LIVE] 3/3 — execute SQL transaction")
    subprocess.run(["scp", "-q", str(sql_path), f"{SSH_HOST}:{SERVER_STAGE}/providers.sql"], check=True)
    subprocess.run(
        ["ssh", SSH_HOST,
         f"sudo -u postgres psql -d ascortbali -v ON_ERROR_STOP=1 -f {SERVER_STAGE}/providers.sql"],
        check=True,
    )
    print("\n[LIVE] Done. 70 providers + their images now on the live site.")

# ──────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--mode", choices=["dry-run", "live"], default="dry-run")
    args = ap.parse_args()

    uploads, sql_path, gcs_path = build_artifacts()

    if args.mode == "dry-run":
        print("\n[DRY-RUN] Nothing executed. Review the files in import_run/ then re-run with --mode live")
        return 0

    # Final guard before destructive live run
    print("\n[LIVE MODE] About to execute against PRODUCTION (gda-pn01 / GCS / ascortbali).")
    ans = input("Type 'GO LIVE' to proceed: ").strip()
    if ans != "GO LIVE":
        print("Aborted.")
        return 1
    run_live(uploads, sql_path, gcs_path)
    return 0


if __name__ == "__main__":
    sys.exit(main())
